# Practice Questions:
import openpyxl
from openpyxl import Workbook
# For the following questions, imagine you have a Workbook object in the variable wb, a Worksheet object in sheet, and a Cell object in cell.

# 1. What does the openpyxl.load_workbook() function return?
# A workbook

# 2. What does the wb.sheetnames workbook attribute contain?
# List of sheetnames as strings

# 3. How would you retrieve the Worksheet object for a sheet named 'Sheet1'?
# Create a variable with
sheet_name = wb['Sheet1']

# 4. How would you retrieve the Worksheet object for the workbook’s active sheet?
print(wb.active)

# 5. How would you retrieve the value in the cell C5?
cell_value = sheet_name['C5']
print(sheet_name['C5'].value)

# 6. How would you set the value in the cell C5 to "Hello"?
sheet_name['C5'] = 'Hello'
print(sheet_name['C5'].value)

# 7. How would you retrieve the cell’s row and column as integers?
print('row %s, column %s' %(wb.row, wb.column))

# 8. How would you save the workbook to the filename example.xlsx?

# 9. If you needed to get the integer index for column 'M', what function would you need to call?

# 10. If you needed to get the string name for column 14, what function would you need to call?

# BONUS: What do the sheet.max_column and sheet.max_row sheet attributes hold, and what is the data type of these attributes?

# BONUS: How can you retrieve a tuple of all the Cell objects from A1 to F1?

# BONUS: How would you set the height of row 5 to 100?
