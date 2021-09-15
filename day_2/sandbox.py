import os
import openpyxl
# from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook



wb = openpyxl.load_workbook('itp_week_3\day_2\lecture.xlsx')
#print(type(wb.sheetnames))

# my_sheet1 = wb['Population by Census Tract']
# my_column = my_sheet1
# #print(type(my_sheet1))


# for i in range(1, my_sheet1.max_row+1):
#     if i == None:
#         break
#     i = str(i)
#     print(i + "I'm a string")

# Creating new sheets
for i in range(3,10):
    wb.create_sheet()

print(str(wb.sheetnames))