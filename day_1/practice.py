# on A, C, of B was sold

import openpyxl

filename = "itp_week_3\day_1\lecture.xlsx"
wb = openpyxl.load_workbook(filename)
print(type(wb))

sheet = wb['Sheet1']
print(sheet.max_row) # 7

for i in range(1, sheet.max_row + 1):  #"sheet.max_row + 1" makes sure last row is included
    # on the date of A, C amount of B were sold.
    date = "A" + str(i)
    date_cell = sheet[date]

    amount = "C" + str(i)
    amount_cell = sheet[amount]

    fruit = "B" + str(i)
    fruit_cell = sheet[fruit]

    print("On the Date of " + str(date_cell.value) + ", " + str(amount_cell.value) + " amount of " + fruit_cell.value + " were sold!")


# import openpyxl


# filename = "itp_week_3\day_1\lecture.xlsx"
# wb = openpyxl.load_workbook(filename)

# for i in range(1, 8):
#     print(i, my_sheet.cell(row = i, column =2).value)